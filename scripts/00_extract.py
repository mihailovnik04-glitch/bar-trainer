# -*- coding: utf-8 -*-
"""00_extract.py — вытаскивает из исходного .xlsx всё сырьё.

Вход:  data/source.xlsx  (исходное барное пособие)
Выход: data/data.json    — {лист: [[номер_строки, знач_кол1, знач_кол2, ...], ...]}
       data/images.json  — {лист: [{file, row, col, row_to, col_to}, ...]}
       img/imageNN.jpg   — все картинки, ужатые до 900 px и JPEG q80

Запускать только если исходник изменился. Иначе используйте готовые data/*.json.
"""
import json, os, zipfile, glob
from pathlib import Path
from xml.etree import ElementTree as ET
import openpyxl
from PIL import Image

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / 'data' / 'source.xlsx'
NS = {'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
      'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
R_ID = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id'
R_EMBED = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed'
MAIN = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'


def dump_cells():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    out = {}
    for ws in wb.worksheets:
        rows = []
        for r in range(1, ws.max_row + 1):
            row = [str(ws.cell(r, c).value) if ws.cell(r, c).value is not None else ''
                   for c in range(1, min(ws.max_column, 40) + 1)]
            if any(x.strip() for x in row):
                rows.append([r] + row)
        out[ws.title] = rows
        print(f'  {ws.title}: {len(rows)} строк')
    json.dump(out, open(ROOT / 'data' / 'data.json', 'w'), ensure_ascii=False)


def dump_images():
    """Якоря картинок: без них невозможно понять, к какому напитку относится фото."""
    z = zipfile.ZipFile(SRC)
    wbrels = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
    relmap = {r.get('Id'): r.get('Target') for r in wbrels}
    wb = ET.fromstring(z.read('xl/workbook.xml'))
    result = {}
    for s in wb.find(MAIN + 'sheets'):
        name = s.get('name')
        target = relmap[s.get(R_ID)]
        path = 'xl/' + target if not target.startswith('/') else target[1:]
        wsx = ET.fromstring(z.read(path))
        drw = wsx.find(MAIN + 'drawing')
        if drw is None:
            result[name] = []
            continue
        rels = ET.fromstring(z.read(path.replace('worksheets/', 'worksheets/_rels/') + '.rels'))
        dtarget = {r.get('Id'): r.get('Target') for r in rels}[drw.get(R_ID)]
        dpath = 'xl/' + dtarget.replace('../', '')
        dx = ET.fromstring(z.read(dpath))
        drels = {r.get('Id'): r.get('Target') for r in
                 ET.fromstring(z.read(dpath.replace('drawings/', 'drawings/_rels/') + '.rels'))}
        imgs = []
        for anch in dx:
            frm, to, blip = anch.find('xdr:from', NS), anch.find('xdr:to', NS), anch.find('.//a:blip', NS)
            if blip is None or frm is None:
                continue
            imgs.append({'file': drels.get(blip.get(R_EMBED), '').replace('../', ''),
                         'row': int(frm.find('xdr:row', NS).text) + 1,
                         'col': int(frm.find('xdr:col', NS).text) + 1,
                         'row_to': int(to.find('xdr:row', NS).text) + 1 if to is not None else None,
                         'col_to': int(to.find('xdr:col', NS).text) + 1 if to is not None else None})
        imgs.sort(key=lambda i: (i['row'], i['col']))
        result[name] = imgs
        print(f'  {name}: {len(imgs)} картинок')
    json.dump(result, open(ROOT / 'data' / 'images.json', 'w'), ensure_ascii=False, indent=1)


def dump_media():
    tmp = ROOT / 'data' / '_media'
    tmp.mkdir(exist_ok=True)
    with zipfile.ZipFile(SRC) as z:
        for n in z.namelist():
            if n.startswith('xl/media/'):
                z.extract(n, tmp)
    (ROOT / 'img').mkdir(exist_ok=True)
    total = 0
    for f in sorted(glob.glob(str(tmp / 'xl' / 'media' / '*'))):
        if f.endswith('.wdp'):
            continue          # HD Photo — Pillow не читает, в пособии не используются
        try:
            im = Image.open(f)
        except Exception:
            continue
        w, h = im.size
        m = 900 / max(w, h)
        if m < 1:
            im = im.resize((int(w * m), int(h * m)), Image.LANCZOS)
        if im.mode in ('RGBA', 'LA', 'P'):
            im = im.convert('RGBA')
            bg = Image.new('RGB', im.size, (255, 255, 255))
            bg.paste(im, mask=im.split()[-1])
            im = bg
        else:
            im = im.convert('RGB')
        out = ROOT / 'img' / (os.path.splitext(os.path.basename(f))[0] + '.jpg')
        im.save(out, 'JPEG', quality=80, optimize=True)
        total += out.stat().st_size
    print(f'  картинок: {len(list((ROOT / "img").iterdir()))}, {total/1e6:.1f} MB')


if __name__ == '__main__':
    if not SRC.exists():
        raise SystemExit(f'Нет исходника: {SRC}\nПоложите .xlsx туда и переименуйте в source.xlsx')
    print('Ячейки:');  dump_cells()
    print('Якоря картинок:'); dump_images()
    print('Медиа:');   dump_media()
