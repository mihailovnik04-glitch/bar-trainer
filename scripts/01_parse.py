import openpyxl, json, re
from pathlib import Path
ROOT = Path(__file__).resolve().parent.parent, os
from collections import defaultdict

wb = openpyxl.load_workbook('bar.xlsx', data_only=True)
imgmap = json.load(open('imgmap.json'))

def cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if v is None: return ''
    return re.sub(r'[ \t]+', ' ', str(v)).strip()

# merged D anchors: get value of merged region top-left for any row
def build_merge_lookup(ws):
    lookup = {}
    for rng in ws.merged_cells.ranges:
        v = ws.cell(row=rng.min_row, column=rng.min_col).value
        for r in range(rng.min_row, rng.max_row+1):
            for c in range(rng.min_col, rng.max_col+1):
                lookup[(r,c)] = (rng.min_row, rng.min_col)
    return lookup

def parse_sheet(name):
    ws = wb[name]
    maxr = ws.max_row
    # find block starts: A non-empty, not directly continuing previous A row
    a_rows = [r for r in range(1, maxr+1) if cell(ws,r,1)]
    starts = []
    for r in a_rows:
        if r-1 in a_rows and starts:
            continue
        starts.append(r)
    blocks = []
    for i, s in enumerate(starts):
        e = (starts[i+1]-1) if i+1 < len(starts) else maxr
        title_parts = []
        rr = s
        while cell(ws, rr, 1):
            title_parts.append(cell(ws, rr, 1)); rr += 1
        ings = []
        methods = []
        volume = ''
        serve = []
        for r in range(s, e+1):
            b = cell(ws, r, 2); c = cell(ws, r, 3); d = cell(ws, r, 4)
            if b:
                ings.append((b, c))
            elif c:
                volume = c
                if d: serve.append(d)
                d = ''
            if d and d not in methods and d not in serve:
                methods.append(d)
        blocks.append(dict(sheet=name, row=s, endrow=e, title=title_parts,
                           ings=ings, method=methods, volume=volume, serve=serve))
    # attach images
    imgs = imgmap.get(name, [])
    for b in blocks:
        main=[]; extra=[]
        for im in imgs:
            if b['row'] <= im['row'] <= b['endrow']:
                (main if im['col']<=2 else extra).append(im['img'])
        b['img'] = main
        b['img_extra'] = extra
    return blocks

out = {}
for sh in ['Лонг Дринки','Шотики','Лимонады и БА','Горячие','Лимонады с собой (Акция)','Самовывоз','Кофе','ПФ','Спец. подачи','Чай ','Украшения','Трубочки','Сахарный сироп ПФ']:
    out[sh] = parse_sheet(sh)
    print(sh, len(out[sh]))
json.dump(out, open('parsed.json','w'), ensure_ascii=False, indent=1)
